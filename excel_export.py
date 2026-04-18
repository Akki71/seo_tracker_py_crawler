"""
excel_export.py — Generate the Excel report.
Extracted from original seo_crawler.py, adapted for API (no stdin).
"""

import json
from datetime import datetime
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILL_PASS = PatternFill("solid", fgColor="D5F5E3")
FILL_FAIL = PatternFill("solid", fgColor="FADBD8")
FONT_PASS = Font(color="1E8449")
FONT_FAIL = Font(color="C0392B")
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

EXPORT_COLUMNS = [
    "url","url_cleaned","status","redirect_suggestion","redirect_type","redirect_target",
    "canonical_status","canonical_url","duplicate_status","word_count","thin_content",
    "current_title","title_length","current_meta_description","meta_desc_length",
    "current_h1","h2_tags","google_analytics","google_search_console",
    "og_tags","og_title_current","og_description_current",
    "schema_markup","schema_types_found",
    "image_alt_status","primary_keyword","secondary_keywords",
    "short_tail_keywords","long_tail_keywords",
    "ai_meta_title","ai_meta_description","ai_h1",
    "ai_og_title","ai_og_description",
    "ai_schema_recommendation","ai_schema_code_snippet",
    "ai_optimized_url","serp_preview",
    "mobile_score","mobile_lcp","mobile_cls","mobile_fcp",
    "desktop_score","desktop_lcp","desktop_cls","desktop_fcp",
    "seo_score","seo_grade","spam_malware_flags","aeo_faq",
    "viewport_configured","html_size_kb","html_size_issue","is_secure",
    "mixed_content","amp_link","og_validation","x_robots_noindex",
    "page_cache_control","crawl_depth","hreflang_tags",
]

def _hdr(ws, row, col, value, fg="2C3E50"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color="FFFFFF", size=9)
    c.fill = PatternFill("solid", fgColor=fg)
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = border
    return c


def generate_excel(pages, broken_links, images, scorecard_results, global_checks,
                   keyword_data, blog_topics_data, backlink_strategy_data,
                   six_month_plan_data, internal_linking_data,
                   keyword_url_map_data, axo_data,
                   base_url, domain, timestamp,
                   robots_status, sitemap_status, llm_status, gbp_status) -> str:

    excel_file = f"output/{domain}_{timestamp}_SEO.xlsx"

    import os; os.makedirs("output", exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: Executive Summary ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Executive Summary"
    ws["A1"] = "AquilTechLabs SEO Audit Report"
    ws["A1"].font = Font(bold=True, size=18, color="2C3E50")
    ws["A2"] = f"Website: {base_url}"
    ws["A3"] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A4"] = f"Pages Crawled: {len(pages)}"
    ws.append([])

    pages_200 = [p for p in pages if str(p.get("status","")).startswith("200") or p.get("status") == 200]
    pages_404 = [p for p in pages if str(p.get("status","")) == "404" or p.get("status") == 404]
    missing_imgs = [i for i in images if i.get("alt_status") == "Missing"]

    summary_rows = [
        ["Metric", "Value"],
        ["Total Pages Crawled", len(pages)],
        ["Pages with 200 OK", len(pages_200)],
        ["Pages with 404", len(pages_404)],
        ["Broken Links Found", len(broken_links)],
        ["robots.txt", robots_status],
        ["sitemap.xml", sitemap_status],
        ["llms.txt", llm_status],
        ["Total Images Audited", len(images)],
        ["Images Missing ALT Text", len(missing_imgs)],
        ["Duplicate Content Pages", len([p for p in pages if "Duplicate" in str(p.get("duplicate_status",""))])],
        ["Thin Content Pages", len([p for p in pages if p.get("thin_content") == "Yes"])],
        ["Pages Missing OG Tags", len([p for p in pages if p.get("og_tags") == "Missing"])],
        ["Pages Missing Schema", len([p for p in pages if p.get("schema_markup") == "Missing"])],
        ["Google Business Profile", gbp_status],
    ]

    for ri, row in enumerate(summary_rows):
        ws.append(row)
        cr = ws.max_row
        for ci in range(1, 3):
            c = ws.cell(row=cr, column=ci)
            c.border = border
            c.alignment = Alignment(horizontal="left")
            if ri == 0:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="2C3E50")

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25

    # ── Sheet 2: All Pages ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("All Pages")
    for ci, col in enumerate(EXPORT_COLUMNS, 1):
        c = ws2.cell(row=1, column=ci, value=col.replace("_"," ").title())
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor="2C3E50")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border

    for ri, pd in enumerate(pages, 2):
        for ci, col in enumerate(EXPORT_COLUMNS, 1):
            val = pd.get(col, "")
            c = ws2.cell(row=ri, column=ci, value=str(val) if val != "" else "")
            c.border = border
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if col in ("url","url_cleaned","canonical_url") and str(val).startswith("http"):
                c.hyperlink = str(val)
                c.font = Font(color="2980B9", underline="single")

    for ci in range(1, len(EXPORT_COLUMNS)+1):
        ws2.column_dimensions[get_column_letter(ci)].width = min(30, 14)
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Broken Links ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Broken Links")
    for ci, h in enumerate(["Source Page","Broken URL","Status","Redirect Suggestion"], 1):
        _hdr(ws3, 1, ci, h, "C0392B")
    for ri, bl in enumerate(broken_links, 2):
        ws3.cell(row=ri, column=1, value=bl.get("source_page","")).border = border
        ws3.cell(row=ri, column=2, value=bl.get("broken_url","")).border = border
        ws3.cell(row=ri, column=3, value=str(bl.get("status",""))).border = border
        ws3.cell(row=ri, column=4, value=bl.get("redirect_suggestion","")).border = border
    for col in ['A','B','C','D']:
        ws3.column_dimensions[col].width = 45

    # ── Sheet 4: Image Audit ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("Image Audit")
    for ci, h in enumerate(["Page URL","Image Source","ALT Text","ALT Status","AI Recommended ALT"], 1):
        _hdr(ws4, 1, ci, h, "8E44AD")
    for ri, img in enumerate(images, 2):
        ws4.cell(row=ri, column=1, value=img.get("page","")).border = border
        src_cell = ws4.cell(row=ri, column=2, value=img.get("src",""))
        src_cell.border = border
        if img.get("src","").startswith("http"):
            src_cell.hyperlink = img["src"]
            src_cell.font = Font(color="2980B9", underline="single")
        ws4.cell(row=ri, column=3, value=img.get("alt","")).border = border
        c = ws4.cell(row=ri, column=4, value=img.get("alt_status",""))
        c.border = border
        if img.get("alt_status") == "Missing":
            c.fill = PatternFill("solid", fgColor="FADBD8")
        ws4.cell(row=ri, column=5, value=img.get("ai_alt_recommendation","")).border = border
    for col,w in zip(['A','B','C','D','E'],[45,60,35,15,50]):
        ws4.column_dimensions[col].width = w

    # ── Sheet 5: Scorecard ─────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Scorecard")
    for ci, h in enumerate(["SEO Parameter","Pass","Fail","Total","Pass %","Status"], 1):
        _hdr(ws5, 1, ci, h)
    for ri, (label, pc, fc, tot, pct, status) in enumerate(scorecard_results, 2):
        ws5.cell(row=ri, column=1, value=label).border = border
        ws5.cell(row=ri, column=2, value=pc).border = border
        ws5.cell(row=ri, column=3, value=fc).border = border
        ws5.cell(row=ri, column=4, value=tot).border = border
        ws5.cell(row=ri, column=5, value=f"{pct:.1f}%").border = border
        sc = ws5.cell(row=ri, column=6, value=status)
        sc.border = border
        sc.font = Font(bold=True)
        if status == "PASSED":
            sc.fill = FILL_PASS; sc.font = Font(bold=True, color="1E8449")
        elif status == "FAILED":
            sc.fill = FILL_FAIL; sc.font = Font(bold=True, color="C0392B")
        else:
            sc.fill = PatternFill("solid", fgColor="FCF3CF"); sc.font = Font(bold=True, color="B7950B")
    for col, w in zip(['A','B','C','D','E','F'],[40,12,12,12,12,12]):
        ws5.column_dimensions[col].width = w

    # ── Sheet 6: Keywords ──────────────────────────────────────────────────────
    if keyword_data.get("services"):
        ws6 = wb.create_sheet("SEO Keywords")
        ws6["A1"] = f"SEO Keyword Analysis — {domain}"
        ws6["A1"].font = Font(bold=True, size=14, color="1F4E79")
        for ci, h in enumerate(["Service","Keyword","Primary","Secondary","Short Tail","Long Tail"], 1):
            _hdr(ws6, 3, ci, h, "1F4E79")
        kw_row = 4
        for svc in keyword_data["services"]:
            for kw in svc.get("keywords", []):
                ws6.cell(row=kw_row, column=1, value=svc["service"]).border = border
                ws6.cell(row=kw_row, column=2, value=kw).border = border
                ws6.cell(row=kw_row, column=3, value=svc.get("primary","")).border = border
                ws6.cell(row=kw_row, column=4, value=", ".join(svc.get("secondary",[]))).border = border
                ws6.cell(row=kw_row, column=5, value=", ".join(svc.get("short_tail",[]))).border = border
                ws6.cell(row=kw_row, column=6, value=", ".join(svc.get("long_tail",[]))).border = border
                kw_row += 1
        for col,w in zip(['A','B','C','D','E','F'],[25,35,30,45,28,50]):
            ws6.column_dimensions[col].width = w

    # ── Sheet 7: Blog Topics ───────────────────────────────────────────────────
    if blog_topics_data:
        ws7 = wb.create_sheet("Blog Topics")
        ws7["A1"] = f"Blog Topics — {domain}"
        ws7["A1"].font = Font(bold=True, size=14, color="27AE60")
        for ci, h in enumerate(["Service","Title","Type","Target Keyword","Description"], 1):
            _hdr(ws7, 3, ci, h, "27AE60")
        bt_row = 4
        for svc in blog_topics_data:
            for t in svc.get("topics", []):
                ws7.cell(row=bt_row, column=1, value=svc.get("service","")).border = border
                ws7.cell(row=bt_row, column=2, value=t.get("title","")).border = border
                ws7.cell(row=bt_row, column=3, value=t.get("type","")).border = border
                ws7.cell(row=bt_row, column=4, value=t.get("target_keyword","")).border = border
                ws7.cell(row=bt_row, column=5, value=t.get("description","")).border = border
                bt_row += 1
        for col,w in zip(['A','B','C','D','E'],[25,50,15,35,60]):
            ws7.column_dimensions[col].width = w

    wb.save(excel_file)
    return excel_file